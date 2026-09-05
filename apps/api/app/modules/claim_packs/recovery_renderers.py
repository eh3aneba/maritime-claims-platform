from __future__ import annotations

from io import BytesIO
import textwrap
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from pypdf import PdfReader, PdfWriter

from app.modules.claim_packs.renderers import render_pdf as _render_base_pdf
from app.modules.claim_packs.renderers import render_xlsx as _render_base_xlsx


def _safe(value: Any) -> str:
    if value is None:
        return "Not established"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value)


def _assessment_lines(snapshot: dict[str, Any]) -> list[str]:
    assessment = snapshot.get("approved_assessment")
    lines = ["APPROVED INITIAL ASSESSMENT — IMMUTABLE AUDIT HANDOFF"]
    if not assessment:
        lines.append("No digest-bound approved Initial Assessment is included in this Claim Pack.")
        return lines
    lines.extend(
        [
            assessment.get("disclaimer", "Approved assessment reporting context only."),
            "",
            f"Assessment version: v{assessment.get('version')}",
            f"Classification: {_safe(assessment.get('classification'))}",
            f"Status: {_safe(assessment.get('status'))}",
            f"Approved at: {_safe(assessment.get('approved_at'))}",
            f"Source state at export: {_safe(assessment.get('source_state_at_export'))}",
            f"Bound source fingerprint: {_safe(assessment.get('source_fingerprint'))}",
            f"Approved content digest: {_safe(assessment.get('approved_content_hash'))}",
            "",
            "The digest above identifies the persisted human-approved assessment. Later source evolution may mark the "
            "historical version stale but never rewrites its approved content or digest.",
        ]
    )
    return lines


def _recovery_lines(snapshot: dict[str, Any]) -> list[str]:
    recovery = snapshot.get("recovery_review") or {}
    summary = recovery.get("summary") or {}
    lines = [
        "RECOVERY / TIME-BAR HUMAN REVIEW",
        recovery.get("disclaimer", "Recovery review not available."),
        "",
        f"Human closure review state: {_safe(recovery.get('human_closure_review_state'))}",
        f"Current counterparties: {_safe(summary.get('counterparty_count', 0))}",
        f"Current time-bar scenarios: {_safe(summary.get('timebar_scenario_count', 0))}",
        f"Current human decisions: {_safe(summary.get('human_decision_count', 0))}",
        f"Recorded human actions: {_safe(summary.get('human_action_count', 0))}",
        f"Open pursue/monitor decisions: {_safe(summary.get('open_human_decision_count', 0))}",
        f"Stale human decisions: {_safe(summary.get('stale_human_decision_count', 0))}",
        f"Unreviewed counterparties: {_safe(summary.get('unreviewed_counterparty_count', 0))}",
        f"Stale time-bar scenarios: {_safe(summary.get('stale_timebar_scenario_count', 0))}",
        f"Unreviewed time-bar scenarios: {_safe(summary.get('unreviewed_timebar_scenario_count', 0))}",
        "",
        "CLOSURE REVIEW BLOCKERS",
    ]
    blockers = recovery.get("closure_review_blockers") or []
    if blockers:
        lines.extend(f"- {item}" for item in blockers)
    else:
        lines.append("No open recovery-path blocker is recorded by this projection. Human closure authority remains unchanged.")

    lines.extend(["", "CURRENT HUMAN RECOVERY DISPOSITIONS"])
    decisions = recovery.get("decisions") or []
    if not decisions:
        lines.append("No explicit human recovery disposition is recorded.")
    for decision in decisions:
        lines.append(
            f"{decision.get('counterparty_name')} [{decision.get('counterparty_role')}]: "
            f"{decision.get('disposition')} / context {decision.get('context_state_status')} / v{decision.get('version')}"
        )
        lines.append(f"  Basis: {decision.get('basis_reference')}")
        lines.append(f"  Rationale: {decision.get('rationale')}")
        if decision.get("next_review_date"):
            lines.append(f"  Next human review: {decision.get('next_review_date')}")
        for action in reversed(decision.get("actions") or []):
            lines.append(
                f"  Action #{action.get('action_number')} {action.get('occurred_on')} "
                f"[{action.get('action_type')}/{action.get('direction')}]: {action.get('summary')}"
            )
            lines.append(f"    Source: {action.get('source_reference')}")

    lines.extend(["", "CURRENT TIME-BAR REVIEW CONTEXT"])
    scenarios = recovery.get("timebar_scenarios") or []
    if not scenarios:
        lines.append("No human-defined time-bar scenario is recorded.")
    for scenario in scenarios:
        review = scenario.get("latest_review")
        review_text = (
            f"human review {review.get('action')} / confirmed {review.get('confirmed_deadline') or 'not established'}"
            if review
            else "no human/legal review recorded"
        )
        lines.append(
            f"{scenario.get('title')}: candidate {scenario.get('candidate_deadline')} / "
            f"source {scenario.get('source_state_status')} / {review_text}"
        )
        lines.append(f"  Legal basis entered by human: {scenario.get('legal_basis')}")
        lines.append(f"  Source reference: {scenario.get('source_reference')}")
    return lines


def _appendix_lines(snapshot: dict[str, Any]) -> list[str]:
    lines = _assessment_lines(snapshot) + ["", ""] + _recovery_lines(snapshot)
    wrapped: list[str] = []
    for line in lines:
        if not line:
            wrapped.append("")
        else:
            wrapped.extend(textwrap.wrap(str(line), width=94, replace_whitespace=False) or [""])
    return wrapped


def _pdf_safe(value: str) -> str:
    return value.encode("cp1252", "replace").decode("cp1252").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _text_pdf(lines: list[str]) -> bytes:
    pages = [lines[index : index + 54] for index in range(0, len(lines), 54)] or [[]]
    objects: list[bytes] = []
    page_numbers = [4 + index * 2 for index in range(len(pages))]
    content_numbers = [5 + index * 2 for index in range(len(pages))]
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    kids = " ".join(f"{number} 0 R" for number in page_numbers)
    objects.append(f"<< /Type /Pages /Kids [{kids}] /Count {len(pages)} >>".encode("ascii"))
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    for page_index, (page_lines, content_number) in enumerate(zip(pages, content_numbers, strict=True), start=1):
        page_obj = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] "
            f"/Resources << /Font << /F1 3 0 R >> >> /Contents {content_number} 0 R >>"
        ).encode("ascii")
        commands = ["BT", "/F1 9 Tf", "54 790 Td", "12 TL"]
        for line in page_lines:
            commands.append(f"({_pdf_safe(line)}) Tj")
            commands.append("T*")
        commands.extend(["ET", "BT", "/F1 8 Tf", "54 28 Td", f"(Governed review appendix - page {page_index} of {len(pages)}) Tj", "ET"])
        stream = "\n".join(commands).encode("cp1252", "replace")
        content_obj = f"<< /Length {len(stream)} >>\nstream\n".encode("ascii") + stream + b"\nendstream"
        objects.extend([page_obj, content_obj])
    output = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for number, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{number} 0 obj\n".encode("ascii"))
        output.extend(obj)
        output.extend(b"\nendobj\n")
    xref = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend((f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n").encode("ascii"))
    return bytes(output)


def render_pdf(snapshot: dict[str, Any]) -> bytes:
    base = PdfReader(BytesIO(_render_base_pdf(snapshot)))
    appendix = PdfReader(BytesIO(_text_pdf(_appendix_lines(snapshot))))
    writer = PdfWriter()
    for page in base.pages:
        writer.add_page(page)
    for page in appendix.pages:
        writer.add_page(page)
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def _prepare_assessment_sheet(workbook, snapshot: dict[str, Any]) -> None:
    if "Approved Assessment" not in workbook.sheetnames:
        return
    sheet = workbook["Approved Assessment"]
    sheet.insert_rows(1, amount=9)
    assessment = snapshot.get("approved_assessment")
    rows = [
        ("Control", "Digest-bound approved assessment handoff only"),
        ("Authority", assessment.get("authority") if assessment else "No eligible approved assessment"),
        ("Version", f"v{assessment.get('version')}" if assessment else "None"),
        ("Classification", assessment.get("classification") if assessment else "None"),
        ("Source state at export", assessment.get("source_state_at_export") if assessment else "None"),
        ("Source fingerprint", assessment.get("source_fingerprint") if assessment else "None"),
        ("Approved content digest", assessment.get("approved_content_hash") if assessment else "None"),
        ("Control notice", assessment.get("disclaimer") if assessment else "Draft, under-review and undigested legacy assessments are excluded."),
    ]
    for index, row in enumerate(rows, start=1):
        sheet.cell(index, 1, row[0])
        sheet.cell(index, 2, row[1])
        sheet.cell(index, 1).font = Font(bold=True)
    sheet.freeze_panes = "A11"
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def render_xlsx(snapshot: dict[str, Any]) -> bytes:
    workbook = load_workbook(BytesIO(_render_base_xlsx(snapshot)))
    _prepare_assessment_sheet(workbook, snapshot)
    if "Recovery Review" in workbook.sheetnames:
        del workbook["Recovery Review"]
    sheet = workbook.create_sheet("Recovery Review")
    sheet.append(["Kind", "Counterparty / scenario", "State", "Detail", "Source / basis"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    recovery = snapshot.get("recovery_review") or {}
    sheet.append([
        "control",
        "Human closure review",
        recovery.get("human_closure_review_state"),
        recovery.get("disclaimer"),
        "",
    ])
    for blocker in recovery.get("closure_review_blockers") or []:
        sheet.append(["blocker", "", "open", blocker, ""])
    for decision in recovery.get("decisions") or []:
        sheet.append([
            "human decision",
            decision.get("counterparty_name"),
            f"{decision.get('disposition')} / {decision.get('context_state_status')}",
            decision.get("rationale"),
            decision.get("basis_reference"),
        ])
        for action in reversed(decision.get("actions") or []):
            sheet.append([
                f"action #{action.get('action_number')}",
                decision.get("counterparty_name"),
                f"{action.get('action_type')} / {action.get('direction')}",
                action.get("summary"),
                action.get("source_reference"),
            ])
    for scenario in recovery.get("timebar_scenarios") or []:
        review = scenario.get("latest_review") or {}
        sheet.append([
            "time-bar scenario",
            scenario.get("title"),
            f"source {scenario.get('source_state_status')} / review {review.get('action') or 'none'}",
            f"Candidate deadline {scenario.get('candidate_deadline')}; human confirmed {review.get('confirmed_deadline') or 'not established'}",
            f"{scenario.get('legal_basis')} | {scenario.get('source_reference')}",
        ])
    widths = [22, 36, 34, 90, 70]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
