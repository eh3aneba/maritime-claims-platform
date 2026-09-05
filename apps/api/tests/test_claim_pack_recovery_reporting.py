from __future__ import annotations

from datetime import date
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook
from pypdf import PdfReader

from app.modules.claim_packs.recovery_renderers import render_pdf, render_xlsx
from app.modules.claim_packs.recovery_snapshot import build_recovery_snapshot
from app.modules.claim_packs.service import _jsonable, _snapshot_hash


def _counterparty(identifier: str, *, state: str = "reference_only") -> dict:
    return {
        "id": identifier,
        "counterparty_key": "counterparty",
        "version": 1,
        "supersedes_id": None,
        "created_by_id": None,
        "name": f"Workshop {identifier}",
        "role": "Potential contractor",
        "allegation_basis": "Human hypothesis only",
        "source_reference": "Reviewed correspondence",
        "source_document_id": None,
        "source_document_family_id": None,
        "source_document_version": None,
        "source_document_hash": None,
        "source_state_status": state,
        "record_hash": "a" * 64,
        "created_at": "2026-09-05T00:00:00Z",
    }


def _decision(identifier: str, *, disposition: str, state: str = "reference_only") -> dict:
    return {
        "id": f"decision-{identifier}",
        "decision_key": "decision",
        "version": 1,
        "supersedes_id": None,
        "counterparty_id": identifier,
        "counterparty_name": f"Workshop {identifier}",
        "counterparty_role": "Potential contractor",
        "decided_by_id": None,
        "disposition": disposition,
        "rationale": "Human rationale for recovery handling.",
        "basis_reference": "Recovery review note",
        "next_review_date": date(2026, 9, 30),
        "previous_decision_hash": None,
        "decision_hash": "b" * 64,
        "context_state_status": state,
        "decided_at": "2026-09-05T00:00:00Z",
        "actions": [
            {
                "id": "action-1",
                "decision_key": "decision",
                "decision_id": f"decision-{identifier}",
                "created_by_id": None,
                "action_number": 1,
                "action_type": "correspondence",
                "direction": "outbound",
                "occurred_on": date(2026, 9, 5),
                "summary": "Human-approved preservation correspondence recorded.",
                "source_reference": "REC-001",
                "external_status": None,
                "external_response_date": None,
                "previous_action_hash": None,
                "action_hash": "c" * 64,
                "created_at": "2026-09-05T00:00:00Z",
            }
        ],
    }


def _scenario(*, state: str = "reference_only", reviewed: bool = True) -> dict:
    return {
        "id": "scenario-1",
        "scenario_key": "scenario",
        "version": 1,
        "supersedes_id": None,
        "created_by_id": None,
        "counterparty_id": "cp-1",
        "title": "Contractual recovery scenario",
        "legal_basis": "Human-entered contractual basis",
        "source_reference": "Contract clause review",
        "source_document_id": None,
        "source_document_family_id": None,
        "source_document_version": None,
        "source_document_hash": None,
        "source_state_status": state,
        "anchor_date": date(2026, 1, 1),
        "period_value": 12,
        "period_unit": "months",
        "extension_value": None,
        "extension_unit": None,
        "extension_basis": None,
        "assumptions": "Human assumptions",
        "candidate_deadline": date(2027, 1, 1),
        "scenario_hash": "d" * 64,
        "created_at": "2026-09-05T00:00:00Z",
        "latest_review": (
            {
                "id": "review-1",
                "scenario_id": "scenario-1",
                "reviewed_by_id": None,
                "scenario_hash": "d" * 64,
                "review_number": 1,
                "action": "confirm",
                "confirmed_deadline": date(2027, 1, 1),
                "note": "Human/legal review",
                "source_reference": "Legal review note",
                "previous_review_hash": None,
                "review_hash": "e" * 64,
                "reviewed_at": "2026-09-05T00:00:00Z",
            }
            if reviewed
            else None
        ),
    }


def _build_with(*, disposition: str, decision_state: str = "reference_only", scenario_state: str = "reference_only", scenario_reviewed: bool = True):
    cp = object()
    decision = object()
    scenario = object()
    with (
        patch("app.modules.claim_packs.recovery_snapshot.current_counterparties", return_value=[cp]),
        patch("app.modules.claim_packs.recovery_snapshot.counterparty_response", return_value=_counterparty("cp-1")),
        patch("app.modules.claim_packs.recovery_snapshot.current_decisions", return_value=[decision]),
        patch(
            "app.modules.claim_packs.recovery_snapshot.decision_response",
            return_value=_decision("cp-1", disposition=disposition, state=decision_state),
        ),
        patch("app.modules.claim_packs.recovery_snapshot.current_scenarios", return_value=[scenario]),
        patch(
            "app.modules.claim_packs.recovery_snapshot.scenario_response",
            return_value=_scenario(state=scenario_state, reviewed=scenario_reviewed),
        ),
    ):
        return build_recovery_snapshot(object(), claim=object())


def test_recovery_projection_preserves_open_human_authority() -> None:
    snapshot = _build_with(disposition="monitor")
    assert snapshot["authority"] == "downstream_human_record_projection_only"
    assert snapshot["human_closure_review_state"] == "open_recovery_paths"
    assert snapshot["summary"]["open_human_decision_count"] == 1
    assert snapshot["summary"]["human_action_count"] == 1
    assert any("pursue/monitor" in item for item in snapshot["closure_review_blockers"])


def test_recovery_projection_escalates_stale_context() -> None:
    snapshot = _build_with(disposition="close", decision_state="stale")
    assert snapshot["human_closure_review_state"] == "attention_required"
    assert snapshot["summary"]["stale_human_decision_count"] == 1
    assert any("stale or unavailable context" in item for item in snapshot["closure_review_blockers"])


def test_recovery_projection_reports_terminal_human_path_without_auto_closure() -> None:
    snapshot = _build_with(disposition="do_not_pursue")
    assert snapshot["human_closure_review_state"] == "no_open_recovery_path_recorded"
    assert snapshot["summary"]["terminal_human_decision_count"] == 1
    assert "human handler remains responsible" in snapshot["disclaimer"].lower()


def test_recovery_projection_dates_are_json_safe_before_snapshot_hashing() -> None:
    snapshot = _jsonable({"recovery_review": _build_with(disposition="monitor")})
    decision = snapshot["recovery_review"]["decisions"][0]
    scenario = snapshot["recovery_review"]["timebar_scenarios"][0]
    assert decision["next_review_date"] == "2026-09-30"
    assert decision["actions"][0]["occurred_on"] == "2026-09-05"
    assert scenario["candidate_deadline"] == "2027-01-01"
    assert len(_snapshot_hash(snapshot)) == 64


def _export_snapshot() -> dict:
    recovery = _jsonable(_build_with(disposition="monitor"))
    return {
        "snapshot_schema_version": "1.2",
        "generated_at": "2026-09-05T00:00:00+00:00",
        "generated_by": {"full_name": "Claims Manager"},
        "claim": {
            "claim_reference": "MCRI-HM-TEST",
            "vessel_name": "MT ORION",
            "imo_number": "9999999",
            "status": "open",
            "incident_date": "2026-01-01",
        },
        "summary": {
            "review_state": "reviewed_with_open_items",
            "approved_fact_count": 0,
            "open_conflict_count": 0,
            "outstanding_requirement_count": 0,
            "open_task_count": 0,
            "open_financial_flag_count": 0,
            "approved_assessment_version": None,
        },
        "evidence_matrix": {"rows": []},
        "policy_intelligence": {"disclaimer": "Review only", "terms": [], "issue_spots": []},
        "outstanding_requirements": [],
        "open_tasks": [],
        "financial": {"totals_by_currency": {}, "open_flags": [], "items": []},
        "approved_assessment": None,
        "technical_investigation": {"authority": "human_investigation_review_only", "disclaimer": "Review only", "topics": []},
        "recovery_review": recovery,
    }


def test_xlsx_export_contains_recovery_review_sheet() -> None:
    workbook = load_workbook(BytesIO(render_xlsx(_export_snapshot())))
    assert "Recovery Review" in workbook.sheetnames
    sheet = workbook["Recovery Review"]
    values = [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None]
    assert any("Human closure review" in value for value in values)
    assert any("monitor" in value for value in values)
    assert any("REC-001" in value for value in values)


def test_pdf_export_appends_recovery_review_pages() -> None:
    snapshot = _export_snapshot()
    base_pages = len(PdfReader(BytesIO(__import__("app.modules.claim_packs.renderers", fromlist=["render_pdf"]).render_pdf(snapshot))).pages)
    rendered = PdfReader(BytesIO(render_pdf(snapshot)))
    assert len(rendered.pages) > base_pages
    text = "\n".join(page.extract_text() or "" for page in rendered.pages)
    assert "RECOVERY / TIME-BAR HUMAN REVIEW" in text
    assert "Human closure review state: open_recovery_paths" in text
