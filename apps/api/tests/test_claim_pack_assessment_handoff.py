from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.modules.assessments.service import approve_assessment, generate_assessment, get_assessment, review_section
from app.modules.claim_packs.recovery_renderers import render_xlsx
from app.modules.claim_packs.recovery_service import build_claim_pack_snapshot
from app.modules.claims.models import Claim
from app.modules.users.models import User
from tests.db_harness import TestingSessionLocal, reset_database
from tests.test_initial_assessment import seed


def setup_function() -> None:
    reset_database()


def _snapshot(db, claim, user):
    return build_claim_pack_snapshot(
        db,
        claim=claim,
        user=user,
        generation_note="Assessment handoff regression",
        generated_at=datetime(2026, 9, 5, 10, 0, tzinfo=UTC),
    )


def _approve_current(db, claim, user):
    assessment = generate_assessment(
        db,
        claim=claim,
        user=user,
        allow_if_not_ready=True,
        override_reason="Controlled preliminary assessment for Claim Pack handoff",
    )
    assessment, sections = get_assessment(db, claim=claim, assessment_id=assessment.id)
    for section in sections:
        review_section(
            db,
            claim=claim,
            section=section,
            user=user,
            action="approve",
            text=None,
            expected_source_fingerprint=assessment.source_fingerprint,
        )
    approve_assessment(
        db,
        claim=claim,
        assessment=assessment,
        user=user,
        note="Human-approved handoff snapshot",
        expected_source_fingerprint=assessment.source_fingerprint,
    )
    return assessment


def test_claim_pack_uses_only_digest_bound_approved_assessment_and_preserves_stale_history() -> None:
    claim_id, user_id = seed()
    with TestingSessionLocal() as db:
        claim = db.get(Claim, claim_id)
        user = db.get(User, user_id)

        draft = generate_assessment(
            db,
            claim=claim,
            user=user,
            allow_if_not_ready=True,
            override_reason="Draft must not flow downstream",
        )
        draft_snapshot = _snapshot(db, claim, user)
        assert draft_snapshot["snapshot_schema_version"] == "1.3"
        assert draft_snapshot["approved_assessment"] is None
        assert draft_snapshot["summary"]["approved_assessment_content_hash"] is None

        # Deliberately generate a fresh version because the first draft remains a historical draft.
        approved = _approve_current(db, claim, user)
        assert approved.version == draft.version + 1
        digest = approved.approved_content_hash
        assert digest and len(digest) == 64

        current_snapshot = _snapshot(db, claim, user)
        handoff = current_snapshot["approved_assessment"]
        assert handoff["authority"] == "downstream_approved_assessment_context_only"
        assert handoff["status"] == "approved"
        assert handoff["approved_content_hash"] == digest
        assert handoff["source_fingerprint"] == approved.source_fingerprint
        assert handoff["source_state_at_export"] == "current"
        assert handoff["sections"]

        claim.incident_description = "Turbocharger failure with later reviewed source-state update"
        db.flush()
        stale_snapshot = _snapshot(db, claim, user)
        stale_handoff = stale_snapshot["approved_assessment"]
        assert stale_handoff["source_state_at_export"] == "stale"
        assert stale_handoff["approved_content_hash"] == digest
        assert stale_handoff["source_fingerprint"] == approved.source_fingerprint

        payload = render_xlsx(stale_snapshot)
        workbook = load_workbook(BytesIO(payload), read_only=True)
        sheet = workbook["Approved Assessment"]
        values = "\n".join(
            str(cell.value)
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        assert "Digest-bound approved assessment handoff only" in values
        assert "Approved content digest" in values
        assert digest in values
        assert "Source state at export" in values
        assert "stale" in values


def test_approved_row_without_digest_is_excluded_instead_of_fabricating_integrity_metadata() -> None:
    claim_id, user_id = seed()
    with TestingSessionLocal() as db:
        claim = db.get(Claim, claim_id)
        user = db.get(User, user_id)
        approved = _approve_current(db, claim, user)
        approved.approved_content_hash = None
        db.flush()

        snapshot = _snapshot(db, claim, user)
        assert snapshot["approved_assessment"] is None
        assert snapshot["summary"]["approved_assessment_version"] is None
        assert snapshot["summary"]["approved_assessment_content_hash"] is None
